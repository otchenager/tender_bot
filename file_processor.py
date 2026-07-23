"""
AI agent: 4-step pipeline for tender analysis.

Step 1 — Claude extracts smeta positions from documents.
Step 2 — Claude matches positions against contractor price list.
Step 3 — Python calculates K, L, M, S scores (never Claude).
Step 4 — Claude writes a 2-3 sentence Russian comment.
"""

import base64
import io
import json
import re
import time
from datetime import datetime

import anthropic
import requests
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

import db
import notifier
from config import ANTHROPIC_API_KEY, ANTHROPIC_MODEL
from logger import get_logger

log = get_logger("file_processor")

_client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY, max_retries=3)

_CATEGORIES = [
    "отделка_потолков", "отделка_стен", "облицовка_стен", "полы",
    "облицовка_полов", "металлоконструкции", "двери_окна",
    "отопление", "электрика", "леса", "прочее",
]

_ADJACENT = {
    "отделка_потолков": {"отделка_стен", "облицовка_стен"},
    "отделка_стен":     {"отделка_потолков", "облицовка_стен"},
    "облицовка_стен":   {"отделка_стен", "облицовка_полов"},
    "полы":             {"облицовка_полов"},
    "облицовка_полов":  {"полы", "облицовка_стен"},
    "металлоконструкции": {"двери_окна", "леса"},
    "двери_окна":       {"металлоконструкции"},
    "отопление":        {"электрика"},
    "электрика":        {"отопление"},
    "леса":             {"металлоконструкции"},
    "прочее":           set(),
}

# ---------------------------------------------------------------------------
# Text extraction helpers
# ---------------------------------------------------------------------------

def _pdf_to_text(file_bytes: bytes) -> tuple[str, bool]:
    """Returns (text, is_scan)."""
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            pages = [p.extract_text() or "" for p in pdf.pages]
        text = "\n".join(pages).strip()
        if len(text) >= 100:
            return text, False
    except Exception as e:
        log.warning(f"pdfplumber error: {e}")
    return "", True


def _docx_to_text(file_bytes: bytes) -> str:
    from docx import Document
    doc = Document(io.BytesIO(file_bytes))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _xlsx_to_text(file_bytes: bytes, max_chars: int = 100_000) -> str:
    """Dump all sheets as tab-separated text. BY customers routinely attach
    the smeta as a spreadsheet (e.g. 'копия заказчик.xlsx' on etrade_3512202,
    while the PDF is a scan) — without this the pipeline saw no positions."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    lines = []
    total = 0
    for ws in wb.worksheets:
        lines.append(f"=== Лист: {ws.title} ===")
        for row in ws.iter_rows(values_only=True):
            cells = ["" if c is None else str(c) for c in row]
            if not any(c.strip() for c in cells):
                continue
            line = "\t".join(cells).rstrip()
            lines.append(line)
            total += len(line)
            if total > max_chars:
                lines.append("[... таблица усечена ...]")
                wb.close()
                return "\n".join(lines)
    wb.close()
    return "\n".join(lines)


def _pdf_to_images_b64(file_bytes: bytes) -> list[str]:
    try:
        import os
        from pdf2image import convert_from_bytes
        # POPPLER_PATH is only needed on Windows dev machines; on Railway
        # (Linux) poppler must be on PATH and the hardcoded C:\ path broke it.
        images = convert_from_bytes(
            file_bytes, dpi=200, fmt="png",
            poppler_path=os.getenv("POPPLER_PATH") or None,
        )
        result = []
        for img in images:
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            result.append(base64.b64encode(buf.getvalue()).decode("ascii"))
        return result
    except Exception as e:
        log.error(f"pdf2image error: {e}")
        return []


def _extract_text_from_documents(documents: list[tuple[str, bytes]]) -> tuple[str, list[str]]:
    """Returns (combined_text, list_of_scan_image_b64)."""
    texts = []
    images = []
    for filename, file_bytes in documents:
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext == "pdf":
            text, is_scan = _pdf_to_text(file_bytes)
            if is_scan:
                images.extend(_pdf_to_images_b64(file_bytes))
            else:
                texts.append(text)
        elif ext == "docx":
            try:
                texts.append(_docx_to_text(file_bytes))
            except Exception as e:
                log.warning(f"docx error {filename}: {e}")
        elif ext in ("xlsx", "xlsm"):
            try:
                texts.append(_xlsx_to_text(file_bytes))
            except Exception as e:
                log.warning(f"xlsx error {filename}: {e}")
    return "\n\n".join(texts), images


# ---------------------------------------------------------------------------
# Claude API calls
# ---------------------------------------------------------------------------

def _call_claude(prompt: str, images_b64: list[str] = None) -> str:
    content = []
    for img in (images_b64 or []):
        content.append({
            "type": "image",
            "source": {"type": "base64", "media_type": "image/png", "data": img},
        })
    content.append({"type": "text", "text": prompt})
    msg = _client.messages.create(
        model=ANTHROPIC_MODEL,
        max_tokens=8192,
        messages=[{"role": "user", "content": content}],
    )
    return "".join(b.text for b in msg.content if b.type == "text")


def _parse_json_response(raw: str):
    raw = raw.strip()
    if raw.startswith("```"):
        raw = re.sub(r"^```[a-zA-Z]*\n?", "", raw)
        raw = re.sub(r"\n?```$", "", raw)
        raw = raw.strip()
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        pass
    for pat in (r"\{[\s\S]*\}", r"\[[\s\S]*\]"):
        m = re.search(pat, raw)
        if m:
            try:
                return json.loads(m.group(0))
            except json.JSONDecodeError:
                pass
    return None


# ---------------------------------------------------------------------------
# Step 1: Extract positions from documents
# ---------------------------------------------------------------------------

_EXTRACT_PROMPT = """\
Ты — эксперт по строительным сметам Республики Беларусь.
Извлеки все рабочие позиции из документа ниже.

Пропусти итоговые строки: ИТОГО, Накладные расходы, Плановая прибыль, ОХР, ОПР и т.п.

Категории: отделка_потолков / отделка_стен / облицовка_стен / полы /
облицовка_полов / металлоконструкции / двери_окна / отопление / электрика / леса / прочее

Верни ТОЛЬКО валидный JSON без пояснений:
{
  "positions": [
    {
      "num": "1",
      "name": "Полное наименование позиции",
      "unit": "ед. изм.",
      "quantity": 0.0,
      "labor_cost": 0.0,
      "material_cost": 0.0,
      "transport_cost": 0.0,
      "total_cost": 0.0,
      "category": "отделка_стен"
    }
  ],
  "total_budget": 0.0
}

ДОКУМЕНТ:
"""


def _step1_extract(text: str, images: list[str], tender_id: int) -> dict | None:
    prompt = _EXTRACT_PROMPT + (text if text else "(см. изображение)")

    def attempt():
        log.info(f"Step1 input for tender {tender_id}: doc_text_length={len(text)}, first_300_chars={text[:300]!r}")
        try:
            raw = _call_claude(prompt, images_b64=images if not text else None)
        except Exception as e:
            log.error(f"Step1 API call failed: {type(e).__name__}: {str(e)}")
            if hasattr(e, "status_code"):
                log.error(f"Status code: {e.status_code}")
            if hasattr(e, "response"):
                log.error(f"Response: {e.response}")
            if hasattr(e, "body"):
                log.error(f"Body: {e.body}")
            return None
        log.info(f"Step1 raw response for tender {tender_id}: {raw[:1000]!r}")
        return _parse_json_response(raw)

    data = attempt()
    if not isinstance(data, dict) or "positions" not in data:
        log.warning("Step1: bad response, retrying in 10s")
        time.sleep(10)
        data = attempt()
    if not isinstance(data, dict) or "positions" not in data:
        return None
    return data


# ---------------------------------------------------------------------------
# Step 2: Match positions against price list
# ---------------------------------------------------------------------------

_MATCH_PROMPT = """\
Сопоставь позиции сметы с позициями прайс-листа подрядчика.

ПРАЙС-ЛИСТ (используй точные id):
{price_list}

ПОЗИЦИИ СМЕТЫ:
{smeta_positions}

Формула уверенности (confidence) — следуй строго:
  ЛЕКСИКА (0–0.6):
    Одинаковые слова             → 0.6
    Однокоренные слова           → 0.3–0.4
    Одна тема, разные слова      → 0.1–0.2
    Нет связи                    → 0.0
  КАТЕГОРИЯ (0–0.3):
    Та же категория              → 0.3
    Смежные категории            → 0.15
    Разные категории             → 0.0
  ЕД. ИЗМ. (0–0.1):
    Совпадают                    → 0.1
    Не совпадают                 → 0.0

  match_status:
    confidence >= 0.75           → green
    confidence 0.40–0.74         → yellow
    confidence < 0.40            → grey (matched_item и matched_item_id = null)

Верни ТОЛЬКО валидный JSON без пояснений:
{{
  "matches": [
    {{
      "smeta_name": "наименование из сметы",
      "matched_item": "наименование из прайса или null",
      "matched_item_id": 123,
      "confidence": 0.55,
      "match_status": "yellow",
      "reasoning": "lexical(0.35) + category(0.3) + unit(0.0) = 0.65"
    }}
  ]
}}
"""


def _step2_match(positions: list[dict], price_items: list[dict]) -> list[dict] | None:
    price_list_text = json.dumps(
        [{"id": p["id"], "name": p["name"], "unit": p["unit"], "category": p["category"]}
         for p in price_items],
        ensure_ascii=False,
    )
    smeta_text = json.dumps(
        [{"name": p["name"], "unit": p["unit"], "category": p["category"]}
         for p in positions],
        ensure_ascii=False,
    )
    prompt = _MATCH_PROMPT.format(
        price_list=price_list_text,
        smeta_positions=smeta_text,
    )

    def attempt():
        try:
            raw = _call_claude(prompt)
        except Exception as e:
            log.error(f"Step2 API call failed: {type(e).__name__}: {str(e)}")
            return None
        data = _parse_json_response(raw)
        if isinstance(data, dict) and isinstance(data.get("matches"), list):
            return data["matches"]
        log.warning(f"Step2: response did not match expected schema. Raw response: {raw!r}")
        return None

    matches = attempt()
    if not isinstance(matches, list):
        log.warning("Step2: bad response, retrying in 10s")
        time.sleep(10)
        matches = attempt()
    if not isinstance(matches, list):
        return None
    return matches


# ---------------------------------------------------------------------------
# Step 3: Python calculates K, L, M, S (never Claude)
# ---------------------------------------------------------------------------

# Formula constants shared by first-time scoring and post-settings-change
# rescoring. K_MIN is the hardcoded "at least 30% of our price list matched"
# floor; CONFIDENCE_FLOOR separates real matches (green/yellow) from grey.
K_MIN = 0.30
CONFIDENCE_FLOOR = 0.40


def compute_scores_from_merged(merged: list[dict], total_budget: float,
                               settings: dict, active_count: int) -> dict:
    """K/L/M/S over already-matched position rows (Iron rule: Python only,
    never Claude). Rows need smeta_total_cost / confidence / match_status /
    margin_byn — exactly what tender_positions stores, so the same function
    re-scores existing tenders after the user changes X/Y without any new
    Claude call.

    Returns {"fail": reason} or the four scores."""
    if total_budget <= 0:
        total_budget = sum((p.get("smeta_total_cost") or 0) for p in merged) or 1

    # K — matched positions / total active price items
    matched_count = sum(
        1 for p in merged if (p.get("confidence") or 0) >= CONFIDENCE_FLOOR
    )
    k_score = matched_count / active_count if active_count else 0.0

    if k_score < K_MIN:
        return {"fail": "failed_K"}

    # L — weighted relevance
    l_numerator = sum(
        (p.get("smeta_total_cost") or 0) * (p.get("confidence") or 0)
        for p in merged if (p.get("confidence") or 0) >= CONFIDENCE_FLOOR
    )
    l_score = l_numerator / total_budget

    x_threshold = float(settings.get("x_threshold", 30)) / 100
    if l_score < x_threshold:
        return {"fail": "failed_L"}

    # M — margin (only green positions)
    margin_sum = sum(
        p["margin_byn"]
        for p in merged
        if p.get("match_status") == "green" and p.get("margin_byn") is not None
    )
    m_score = margin_sum / total_budget

    y_threshold = float(settings.get("y_threshold", 5)) / 100
    if m_score < y_threshold:
        return {"fail": "failed_M"}

    # S — final score
    s_score = m_score * 0.5 + l_score * 0.3 + k_score * 0.2

    return {
        "k_score": k_score,
        "l_score": l_score,
        "m_score": m_score,
        "s_score": s_score,
    }


def _step3_scores(positions: list[dict], matches: list[dict],
                  total_budget: float, settings: dict) -> dict | None:
    """Returns dict with scores and merged position data, or None if filter fails."""
    active_items = db.get_active_price_items()
    if not active_items:
        log.warning("Step3: price list is empty, cannot score")
        return None

    # Build match lookup by smeta_name
    match_by_name = {m["smeta_name"]: m for m in matches}

    merged = []
    for pos in positions:
        m = match_by_name.get(pos["name"]) or {}
        confidence = float(m.get("confidence") or 0)
        match_status = m.get("match_status") or "grey"
        matched_item_id = m.get("matched_item_id") if match_status != "grey" else None

        qty = float(pos.get("quantity") or 1)
        total_cost = float(pos.get("total_cost") or 0)

        margin_byn = None
        if match_status == "green" and matched_item_id:
            item = db.get_price_item(matched_item_id)
            if item:
                my_price = float(item.get("my_price") or 0)
                # margin = what tender pays - what we do it for
                margin_byn = round(my_price * qty - total_cost, 2)

        merged.append({
            "smeta_name": pos["name"],
            "smeta_unit": pos.get("unit", ""),
            "smeta_quantity": qty,
            "smeta_labor_cost": pos.get("labor_cost"),
            "smeta_material_cost": pos.get("material_cost"),
            "smeta_transport_cost": pos.get("transport_cost"),
            "smeta_total_cost": total_cost,
            "category": pos.get("category"),
            "matched_item_id": matched_item_id,
            "confidence": confidence,
            "match_status": match_status,
            "margin_byn": margin_byn,
        })

    result = compute_scores_from_merged(merged, total_budget, settings, len(active_items))
    result["merged"] = merged
    return result


# ---------------------------------------------------------------------------
# Step 4: AI comment
# ---------------------------------------------------------------------------

_COMMENT_PROMPT = """\
Ты — советник белорусского строительного подрядчика.

Тендер: {title}
Бюджет: {budget} BYN
Совпадение K: {k:.1%}
Релевантность L: {l:.1%}
Маржа M: {m:.1%}

Топ-5 самых дорогих позиций:
{top5}

Последние 5 подходящих тендеров для контекста:
{history}

Напиши 2–3 предложения на русском:
— Стоит ли участвовать в этом тендере?
— Где основная возможность для прибыли?
— Есть ли риски?
"""


def _step4_comment(tender_id: int, tender: dict, scores: dict) -> str:
    top5 = db.get_top_positions(tender_id, limit=5)
    history = db.get_last_suitable_tenders(limit=5)
    prompt = _COMMENT_PROMPT.format(
        title=tender.get("title", ""),
        budget=tender.get("budget_byn", 0),
        k=scores.get("k_score", 0),
        l=scores.get("l_score", 0),
        m=scores.get("m_score", 0),
        top5=json.dumps(top5, ensure_ascii=False),
        history=json.dumps(history, ensure_ascii=False),
    )
    try:
        return _call_claude(prompt)
    except Exception as e:
        log.error(f"Step4 comment failed for tender {tender_id}: {e}")
        return "Комментарий недоступен"


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def analyze_tender(tender_id: int, documents: list[tuple[str, bytes]]):
    """
    Full 4-step analysis. Mutates DB state.
    documents = list of (filename, bytes).

    Never raises: any unexpected failure (bad Claude response, parsing bug,
    etc.) is caught here so the ingest endpoint can still return 200 to the
    VPS caller instead of a 500 — the VPS has no business retrying our own
    pipeline bugs.
    """
    try:
        _analyze_tender_impl(tender_id, documents)
    except Exception as e:
        log.error(f"analyze_tender crashed for tender {tender_id}: {type(e).__name__}: {e}")
        db.reject_tender(tender_id, "ai_error")


def _analyze_tender_impl(tender_id: int, documents: list[tuple[str, bytes]]):
    if not documents:
        log.error(f"No documents for tender {tender_id}")
        db.reject_tender(tender_id, "ai_error")
        return

    # Step 1 — extract positions
    text, images = _extract_text_from_documents(documents)
    if not text and not images:
        log.error(f"Could not extract any content from documents for tender {tender_id}")
        db.reject_tender(tender_id, "ai_error")
        return

    # Filter I — keyword match, now a REAL content check instead of a title
    # guess. VPS-side title filtering (removed) could never see a keyword
    # that's absent from the tender title but present in the smeta itself
    # (e.g. "штукатурка" makes up 70% of a smeta whose title never mentions
    # it) — checking the actual extracted document text fixes that blind
    # spot. Runs here, before the expensive Step1 Claude call, specifically
    # to save that cost on tenders that would fail anyway. Scan-only PDFs
    # (text extraction empty, only page images) can't be keyword-checked
    # cheaply, so they skip this check and proceed to Step1 as before.
    if text:
        keywords = db.get_active_price_item_names()
        low_text = text.lower()
        if keywords and not any(kw.lower() in low_text for kw in keywords):
            db.reject_tender(tender_id, "keyword")
            log.info(f"Tender {tender_id} rejected: keyword not found in smeta text")
            return

    extraction = _step1_extract(text, images, tender_id)
    if extraction is None:
        log.error(f"Step1 extraction failed for tender {tender_id}")
        db.reject_tender(tender_id, "ai_error")
        return

    positions = extraction.get("positions", [])
    total_budget = float(extraction.get("total_budget") or 0)
    if not positions:
        log.warning(f"No positions extracted for tender {tender_id}")
        db.reject_tender(tender_id, "ai_error")
        return

    # Step 2 — match against price list
    active_items = db.get_active_price_items()
    matches = _step2_match(positions, active_items)
    if matches is None:
        log.error(f"Step2 matching failed for tender {tender_id}")
        db.reject_tender(tender_id, "ai_error")
        return

    # Step 3 — scores (Python only)
    settings = db.get_search_settings()
    result = _step3_scores(positions, matches, total_budget, settings)

    if result is None:
        db.reject_tender(tender_id, "ai_error")
        return

    # Persist positions regardless of outcome (useful for debug)
    db.save_tender_positions(tender_id, result["merged"])

    # Auto-add grey positions as unknown inactive price items
    for pos in result["merged"]:
        if pos["match_status"] == "grey":
            db.add_unknown_price_item(
                name=pos["smeta_name"],
                unit=pos["smeta_unit"],
                my_price=pos["smeta_total_cost"],
            )

    if "fail" in result:
        db.reject_tender(tender_id, result["fail"])
        log.info(f"Tender {tender_id} rejected: {result['fail']}")
        return

    # Passed all filters — save scores
    db.update_tender_scores(
        tender_id,
        result["k_score"],
        result["l_score"],
        result["m_score"],
        result["s_score"],
    )

    # Step 4 — AI comment (never rejects tender on failure)
    tender = db.get_tender(tender_id)
    comment = _step4_comment(tender_id, tender, result)
    db.update_tender_ai_comment(tender_id, comment)

    db.update_tender_status(tender_id, "suitable")
    log.info(
        f"Tender {tender_id} → suitable "
        f"K={result['k_score']:.2f} L={result['l_score']:.2f} "
        f"M={result['m_score']:.2f} S={result['s_score']:.2f}"
    )

    notifier.send_email(tender_id)


# ---------------------------------------------------------------------------
# Rescore after settings change (Python only, no Claude)
# ---------------------------------------------------------------------------

def rescore_existing_tenders() -> dict:
    """Re-apply the CURRENT search settings to already-analyzed tenders.

    Called when the user saves new parameters so old results never linger
    under stale criteria. No Claude involved: confidence/match_status/
    margin_byn are already stored per tender_positions row, so only the
    Python filters (B/R) and scoring (K/L/M/S) are re-run. Tenders that
    cannot be re-scored (no stored positions) are archived with an explicit
    marker instead of silently keeping statuses computed under old criteria.
    """
    settings = db.get_search_settings()
    active_count = len(db.get_active_price_items())
    marker = f"результаты до изменения от {datetime.now().strftime('%d.%m.%Y')}"
    stats = {"suitable": 0, "rejected": 0, "archived": 0}

    for t in db.get_rescorable_tenders():
        tid = t["id"]
        positions = db.get_tender_positions(tid)
        if not positions:
            db.archive_tender(tid, marker)
            stats["archived"] += 1
            continue

        # Re-check the coarse budget/region filters first (mirrors the
        # VPS-side filters B/R — those only ran under the OLD settings).
        budget = t.get("budget_byn") or 0
        fail = None
        result = None
        if budget < float(settings.get("min_budget") or 0):
            fail = "failed_B"
        elif settings.get("regions") and t.get("region") not in settings["regions"]:
            fail = "failed_R"
        else:
            result = compute_scores_from_merged(positions, budget, settings, active_count)
            fail = result.get("fail")

        if fail:
            db.reject_tender(tid, fail)
            stats["rejected"] += 1
        else:
            db.update_tender_scores(
                tid,
                result["k_score"], result["l_score"],
                result["m_score"], result["s_score"],
            )
            db.update_tender_status(tid, "suitable")
            stats["suitable"] += 1

    log.info(f"Rescore after settings change: {stats}")
    return stats


def preview_pass_counts(trial_settings: dict) -> dict:
    """Dry-run twin of rescore_existing_tenders: how many stored tenders
    would be suitable under TRIAL settings. Read-only, no Claude, nothing
    is written — feeds the live preview on the settings page."""
    active_count = len(db.get_active_price_items())
    total = 0
    passing = 0
    for t in db.get_rescorable_tenders():
        positions = db.get_tender_positions(t["id"])
        if not positions:
            continue
        total += 1
        budget = t.get("budget_byn") or 0
        if budget < float(trial_settings.get("min_budget") or 0):
            continue
        if trial_settings.get("regions") and t.get("region") not in trial_settings["regions"]:
            continue
        result = compute_scores_from_merged(positions, budget, trial_settings, active_count)
        if "fail" not in result:
            passing += 1
    return {"total": total, "passing": passing}


# ---------------------------------------------------------------------------
# Document downloader (used by parser)
# ---------------------------------------------------------------------------

DOWNLOAD_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/124.0.0.0",
}


def download_documents(doc_urls: list[str]) -> list[tuple[str, bytes]]:
    """Download document files. Returns list of (filename, bytes)."""
    results = []
    for url in doc_urls:
        try:
            resp = requests.get(
                url, headers=DOWNLOAD_HEADERS, timeout=60, verify=False
            )
            if resp.status_code == 200:
                filename = url.rsplit("/", 1)[-1].split("?")[0] or "document"
                results.append((filename, resp.content))
            else:
                log.warning(f"Download failed {url}: HTTP {resp.status_code}")
        except Exception as e:
            log.error(f"Download error {url}: {e}")
    return results
